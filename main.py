from typing import Sized
import openpyxl
from openpyxl.styles import *

wb1 = openpyxl.load_workbook("Excel/test.xlsx")

row = 1
column = 1
list = []

for i in range(row, 5):
    for j in range(column, 4):
            #sprint(wb1["Student"].cell(row= i, column= j).value)
            
            list.append(wb1["Student"].cell(row= i, column= j).value)
    print()

for i in range(len(list)):
    print(list[i])

wb2 = openpyxl.load_workbook("Excel/test1.xlsx")

k = 0
for i in range(row, 5):
    for j in range(column, 4):
            wb2["Student"].cell(row= i, column=j, value= list[k])
            k += 1


normal = Side(border_style="thin")
all_border = Border(top= normal, bottom= normal, left= normal, right= normal)
all_alignment = Alignment(horizontal="center", vertical= "center")

header_font = Font(size= 12, bold= True)
header_fill = PatternFill(fill_type= "solid", fgColor= "a8a8a8")

for i in range(1, 4):
    wb2["Student"].cell(row= 1, column= i).font = header_font
    wb2["Student"].cell(row= 1, column= i).fill = header_fill


for i in range(1, 5):
    for j in range(1, 4):
        wb2["Student"].cell(row= i, column= j).border = all_border
        wb2["Student"].cell(row= i, column= j).alignment = all_alignment


wb2.save("Excel/test1.xlsx")

