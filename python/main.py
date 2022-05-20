import re
import openpyxl
import json

from openpyxl.styles import *

class Excel():
    def __init__(self, path: str, sheet: int):
        self.path = path
        self.workbook = openpyxl.load_workbook(self.path, data_only=True)

        wb_sheet = self.workbook.sheetnames

        self.workbook_sheet = self.workbook[wb_sheet[sheet - 1]]


    def check_range(range: any):
        if (type(range) not in (str, list)):
            raise TypeError("Range must be a type of string or list")

        elif (type(range) == str):
            if (range.isalpha() or range.isnumeric()):
                raise TypeError("Range string must be a combination of character and number")

        elif (type(range) == list):
            if (len(range) == 2):
                for i in range:
                    if type(i) not in (str, int):
                        raise TypeError("Range list can only have a type of string and integer for its values")

            else:
                raise TypeError("Range list can only have 2 values")
                

    def convert_range(range: any):
        Excel.check_range(range)

        if (type(range) == str):
            column = Excel.check_and_convert_string_value(''.join(x for x in range if not x.isdigit()))
            row = int(''.join(x for x in range if x.isdigit()))

        elif (type(range) == list):
            if (type(range[0]) == str):
                column = Excel.check_and_convert_string_value(range[0])
            
            elif (type(range[0]) == int):
                column = range[0]

            if (type(range[1]) == str):
                row = Excel.check_and_convert_string_value(range[1])
            
            elif (type(range[1]) == int):
                row = range[1]

        return column, row


    def check_and_convert_string_value(value: any):
        if(type(value) == str):
            value = [ord(x) - 96 for x in value.lower()]

            new_value = 0
            for i in range(len(value)):
                new_value += value[i] * 26**(len(value) - (i + 1))

        return new_value

    
    def attributes_string(list_of_attributes: any):
        attributes_string = ""
        for i, attribute in enumerate(list_of_attributes):
            if(i == 0):
                attributes_string += attribute
            
            else:
                attributes_string += f", {attribute}"

        return attributes_string


    def get_value_singular(self, range: any):
        column, row = Excel.convert_range(range)
        value = self.workbook_sheet.cell(row = row, column = column).value

        return value

        
    def get_value_multiple(self, start_range: any, end_range: any):
        start_column, start_row = Excel.convert_range(start_range)
        end_column, end_row = Excel.convert_range(end_range)
        
        value = []
        for row in range(start_row, end_row + 1):
            for column in range(start_column, end_column + 1):
                temp_value = self.workbook_sheet.cell(row = row, column = column).value
                value.append(temp_value)

        return value


    def get_value_multiple_2d(self, start_range: any, end_range: any):
        start_column, start_row = Excel.convert_range(start_range)
        end_column, end_row = Excel.convert_range(end_range)

        value_array = []
        for row in range(start_row, end_row + 1):
            temp_value_array = []
            for column in range(start_column, end_column + 1):
                temp_value = self.workbook_sheet.cell(row = row, column = column).value
                temp_value_array.append(temp_value)
            
            value_array.append(temp_value_array)

        return value_array


class Student():
    def get_student(student_2d_array):
        student_array = []
        for student_index, student in enumerate(student_2d_array):
            new_student_dict = {
                "id": student_index + 1,
                "name": student[0],
                "is_muslim": Student.get_is_muslim(student),
                "assignment": Student.get_assignment(student)
            }

            student_array.append(new_student_dict)

        
        return student_array


    def get_is_muslim(student):
        is_muslim = True
        for non_muslim_student in (Student.get_non_muslim_student()):
            if(student[0] == non_muslim_student.get("name")):
                is_muslim = False

        
        return is_muslim


    def get_non_muslim_student():
        non_muslim_student_array = [
            {
                "id": 1,
                "roll_number": 12,
                "name": "Keatryn Kezia P. Sihombing"
            },
            {
                "id": 2,
                "roll_number": 24,
                "name": "Nicholas Davin Yang"
            },
            {
                "id": 3,
                "roll_number": 26,
                "name": "O'neil Kerry Laurent"
            },
            {
                "id": 4,
                "roll_number": 35,
                "name": "Yehezkiel Dio Sinolungan"
            }
        ]

        return non_muslim_student_array


    def get_assignment(student):
        assignment_array = []
        for assignment_index, assignment in enumerate(student[1:]):
            if(assignment == None):
                assignment = 0
            
            elif(assignment == "ü"):
                assignment = 2

            elif(assignment == "NON-MUS"):
                assignment = 3

            new_assignment_dict = {
                "id": assignment_index + 1,
                "status": assignment
            }

            assignment_array.append(new_assignment_dict)


        return assignment_array


class Assignment():
    def get_assignment(assignment_2d_array):
        assignment_array = []        
        for month_index, month in enumerate(assignment_2d_array[0]):
            if(type(month) == str):
                month_count = ""
                for char in month:
                    if(char.isdigit()):
                        month_count += char


                assignment_2d_array[0][month_index] = int(month_count)


            if(type(month) != str):
                month = assignment_2d_array[0][month_index - 1]
                assignment_2d_array[0][month_index] = month
        

        for week_index, week in enumerate(assignment_2d_array[1]):
            if(type(week) == str):
                week_count = ""
                for char in week:
                    if(char.isdigit()):
                        week_count += char


                assignment_2d_array[1][week_index] = int(week_count)

            
            if(type(week) != str):
                week = assignment_2d_array[1][week_index - 1]
                assignment_2d_array[1][week_index] = week


        for assignment_index, assignment in enumerate(assignment_2d_array[3]):
            splitted_assignment = re.split('(\d+)', assignment)
            subject = splitted_assignment[0]
            count = int(splitted_assignment[1])

            if(subject == "PAI"):
                is_for_muslim = True

            elif(subject != "PAI"):
                is_for_muslim = False 

            new_assignment_dict = {
                "id": assignment_index + 1,
                "subject": subject,
                "count": count,
                "is_for_muslim": is_for_muslim,
                "month": assignment_2d_array[0][assignment_index],
                "week": assignment_2d_array[1][assignment_index],
                "due_date": None,
                "description": None,
                "picture": None
            }

            assignment_array.append(new_assignment_dict)


        return assignment_array


class Main():
    def main():
        workbook_combined = Excel("./excel/List Tugas gabunggan.xlsx", 1)
        student_2d_array = workbook_combined.get_value_multiple_2d("B4", "AT38")
        assignment_2d_array = workbook_combined.get_value_multiple_2d("C1", "AT4")

        student_array = Student.get_student(student_2d_array)
        with open("json/student.json", "w") as outfile:
            outfile.write(json.dumps(student_array, indent = 4))  

        assignment_array = Assignment.get_assignment(assignment_2d_array)    
        with open("json/assignment.json", "w") as outfile:
            outfile.write(json.dumps(assignment_array, indent = 4))  


if(__name__ == "__main__"):
    Main.main()