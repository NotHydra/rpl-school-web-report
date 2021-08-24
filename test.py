import openpyxl
from openpyxl.utils.cell import range_boundaries

"""
print(wb.active.title)

for i in range(1, total_row + 1):
    for j in range(1, total_column + 1):
        print(wb["Student"].cell(row= i, column= j).value)
    
    print()
"""

wb = openpyxl.load_workbook("Excel/test.xlsx")

total_row = wb["Student"].max_row
total_column = wb["Student"].max_column

student4 = [4, "Joko", 100]
temp_row = total_row + 1
temp_colum = total_column + 1

j = 0
for i in range(1, temp_colum):
    wb["Student"].cell(row= temp_row, column= i, value= "123")
    j += 1

wb.save("Excel/test.xlsx")