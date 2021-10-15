import openpyxl
import math
from files import list_of_students_name, weekly_task_start_column_range, weekly_task_end_column_range

def run_wanted_list():
    wb_gabunggan = openpyxl.load_workbook("Excel\List Tugas Gabunggan.xlsx")

    wb_gabunggan_row_start_range = 3
    wb_gabunggan_row_end_range = weekly_task_end_column_range[len(weekly_task_end_column_range) - 1]
    wb_gabunggan_maximum_assignment_value = wb_gabunggan_row_end_range - wb_gabunggan_row_start_range
    number_of_maximum_assignment_not_done = 5

    # Get The Wanted Student On All Assignment
    k = 0
    list_of_wanted_students_all_assignment = []
    list_of_total_assignment_per_student = []
    for i in range(4, 39):
        for j in range(wb_gabunggan_row_start_range, wb_gabunggan_row_end_range):
            temp_value = wb_gabunggan.active.cell(row = i, column = j).value

            if temp_value != None:
                k += 1
        
        temp_list_value = [k, i - 3,]
        list_of_total_assignment_per_student.append(temp_list_value)

        if (k < wb_gabunggan_maximum_assignment_value - number_of_maximum_assignment_not_done):
            list_of_wanted_students_all_assignment.append(list_of_students_name[i-4])

        k = 0


    # Get The Wanted Student On Latest Weekly Assignment
    k = 0
    list_of_wanted_students_latest_weekly_assignment = []
    for i in range(4, 39):
        for j in range(weekly_task_start_column_range[len(weekly_task_start_column_range) - 1], weekly_task_end_column_range[len(weekly_task_end_column_range) - 1]):
            temp_value = wb_gabunggan.active.cell(row = i, column = j).value
            
            if temp_value != None:
                k += 1

        if k < math.ceil((weekly_task_end_column_range[len(weekly_task_end_column_range) - 1] - weekly_task_start_column_range[len(weekly_task_start_column_range) - 1])/2):
            list_of_wanted_students_latest_weekly_assignment.append(list_of_students_name[i-4])

        k = 0

    print()
    print()
    print(list_of_wanted_students_all_assignment, len(list_of_wanted_students_all_assignment))
    print()
    print(list_of_wanted_students_latest_weekly_assignment, len(list_of_wanted_students_latest_weekly_assignment))
    print()

    # Write To All Assignment Text File
    f = open("Excel\Program\Wanted List\list_of_wanted_students_on_all_assignment.txt", "w")
    f.write("*WANTED* \n")

    for i in range(len(list_of_wanted_students_all_assignment)):
        f.write(f"{i + 1}. {list_of_wanted_students_all_assignment[i]} \n")

    f.write("\n*NOTE:* \n")
    f.write(f"Ini nama siswa - siswa yang tugasnya masih ada deadline yang belum dikerjakan {number_of_maximum_assignment_not_done} ke - atas.")
    f.close()

    # Write To Latest Weekly Assignment Text File
    f = open("Excel\Program\Wanted List\list_of_wanted_students_on_latest_weekly_assignment.txt", "w")
    f.write(f"*List Tugas Minggu ke {len(weekly_task_start_column_range)}* \n")

    for i in range(len(list_of_wanted_students_latest_weekly_assignment)):
        f.write(f"{i + 1}. {list_of_wanted_students_latest_weekly_assignment[i]} \n")

    f.write("\n*NOTE:* \n")
    f.write(f"Berikut nama-nama siswa yang belum/mengerjakan tugas kurang dari {math.ceil((weekly_task_end_column_range[len(weekly_task_end_column_range) - 1] - weekly_task_start_column_range[len(weekly_task_start_column_range) - 1])/2)} Tugas dari {weekly_task_end_column_range[len(weekly_task_end_column_range) - 1] - weekly_task_start_column_range[len(weekly_task_start_column_range) - 1]} Tugas.")
    f.close()

    # Write To Sorted List Of Student Assignment
    list_of_total_assignment_per_student.sort()

    f = open("Excel\Program\Wanted List\sorted_list_of_student_assignment.txt", "w")

    for i in range(len(list_of_total_assignment_per_student)):
        f.write(f"{list_of_students_name[list_of_total_assignment_per_student[i][1] - 1]} ({list_of_total_assignment_per_student[i][1]}), Total Laporan : {list_of_total_assignment_per_student[i][0]} \n\n")

    f.close()