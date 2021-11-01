import openpyxl
import json
from files import weekly_task_start_column_range, weekly_task_end_column_range, monthly_task_start_column_range, monthly_task_end_column_range, list_of_students_name, list_of_non_muslim_students

def run_json_writer():
    wb_combined = openpyxl.load_workbook("Excel/List Tugas Gabunggan.xlsx")

    start_column_range = weekly_task_start_column_range[0]
    end_colum_range = weekly_task_end_column_range[len(weekly_task_end_column_range)-1]

    # For Student JSON
    # Get the array of values of assignment has been done or not 
    list_of_student_assignment_values = []
    for i in range(4, 39):
        temp_array = []
        for j in range(start_column_range, end_colum_range):
            temp_value = wb_combined.active.cell(row = i, column = j).value

            temp_array.append(temp_value)
        

        list_of_student_assignment_values.append(temp_array)

    # Set the variables needed
    list_of_student_array = []
    for i in range(len(list_of_students_name)):
        student_id = i+1
        student_name = list_of_students_name[i]

        student_is_muslim = True
        for j in range(len(list_of_non_muslim_students)):
            if list_of_students_name[i] == list_of_non_muslim_students[j]:
                student_is_muslim = False


        student_total_assignment_done = 0
        for j in range(len(list_of_student_assignment_values[i])):
            if list_of_student_assignment_values[i][j] != None:
                student_total_assignment_done += 1


        student_assignment_done = list_of_student_assignment_values[i]

        list_of_student_array.append([student_id, student_name, student_is_muslim, student_total_assignment_done, student_assignment_done])

    # Write to Student JSON file
    json_file = open("Excel/list_of_student.json", "w")
    json_file.write("[\n")

    for i in range(len(list_of_students_name)):
        json_file.write("\t{\n")
        
        json_file.write(f'\t\t"student_id": {list_of_student_array[i][0]},\n')
        json_file.write(f'\t\t"student_name": "{list_of_student_array[i][1]}",\n')

        if list_of_student_array[i][2] == True:
            json_file.write(f'\t\t"student_is_muslim": true ,\n')

        elif list_of_student_array[i][2] == False:
            json_file.write(f'\t\t"student_is_muslim": false ,\n')

        json_file.write(f'\t\t"student_total_asssignment_done": "{list_of_student_array[i][3]}",\n')

        json_file.write('\t\t"student_assignment_done": [')

        for j in range(len(list_of_student_array[i][4])):
            if list_of_student_array[i][4][j] == None:
                json_file.write('0')
            
            elif list_of_student_array[i][4][j] == "ü":
                json_file.write('2')
            
            elif list_of_student_array[i][4][j] == "NON-MUS":
                json_file.write('3')

            if j != len(list_of_student_array[i][4])-1:
                json_file.write(', ')

        json_file.write(']\n')

        if i != len(list_of_students_name)-1:
            json_file.write("\t},\n")
        
        else:
            json_file.write("\t}\n")


    json_file.write("]")
    json_file.close()

    # For Assignment JSON
    # Get the array of values of assignment has been done or not 
    list_of_assignment_values = []
    for j in range(start_column_range, end_colum_range):
        temp_array = []
        for i in range(4, 39):
            temp_value = wb_combined.active.cell(row = i, column = j).value
            temp_array.append(temp_value)

        list_of_assignment_values.append(temp_array)


    # Get the values of task name
    list_of_assignment_name_values = []
    for i in range(start_column_range, end_colum_range):
        temp_value = wb_combined.active.cell(row = 3, column = i).value
        list_of_assignment_name_values.append(temp_value)


    # Set the variables needed
    list_of_assignment_array = []
    for i in range(len(list_of_assignment_values)):
        assignment_id = i+1
        assignment_lesson_name = list_of_assignment_name_values[i].strip("0123456789")
        assignment_lesson_count = list_of_assignment_name_values[i].strip("ABCDEFGHIJKLMNOPQRSTUVWXYZ")
        assignment_description = None
        assignment_picture = None

        assignment_is_for_muslim = False
        if assignment_lesson_name == "PAI":
            assignment_is_for_muslim = True

        for k in range(len(weekly_task_start_column_range)):
            for j in range(weekly_task_start_column_range[k]-2, weekly_task_end_column_range[k]-2):
                if assignment_id == j:
                    assignment_in_week = k+1
            

        for k in range(len(monthly_task_start_column_range)):
            for j in range(monthly_task_start_column_range[k]-2, monthly_task_end_column_range[k]-2):
                if assignment_id == j:
                    assignment_in_month = k+1

        assignment_has_due_date = None
        
        assignment_total_done = 0
        for j in range(len(list_of_students_name)):
            if list_of_assignment_values[i][j] != None:
                assignment_total_done += 1

        assignment_done = list_of_assignment_values[i]

        list_of_assignment_array.append([assignment_id, assignment_lesson_name, assignment_lesson_count, assignment_description, assignment_picture, assignment_is_for_muslim, assignment_in_week, assignment_in_month, assignment_has_due_date, assignment_total_done, assignment_done])


    # Write to Student JSON file
    json_file = open("Excel/list_of_assignment.json", "w")
    json_file.write("[\n")

    for i in range(end_colum_range-3):
        json_file.write("\t{\n")
        
        json_file.write(f'\t\t"assignment_id": {list_of_assignment_array[i][0]},\n')
        json_file.write(f'\t\t"assignment_lesson_name": "{list_of_assignment_array[i][1]}",\n')
        json_file.write(f'\t\t"assignment_lesson_count": {list_of_assignment_array[i][2]},\n')

        if list_of_assignment_array[i][3] == None:
            json_file.write(f'\t\t"assignment_description": null,\n')
        
        else:
            json_file.write(f'\t\t"assignment_description": {list_of_assignment_array[i][3]},\n')
        
        if list_of_assignment_array[i][4] == None:
            json_file.write(f'\t\t"assignment_picture": null,\n')
        
        else:
            json_file.write(f'\t\t"assignment_picture": {list_of_assignment_array[i][4]},\n')
        
        if list_of_assignment_array[i][5] == True:
            json_file.write(f'\t\t"assignment_is_for_muslim": true,\n')
        
        elif list_of_assignment_array[i][5] == False:
            json_file.write(f'\t\t"assignment_is_for_muslim": false,\n')
        
        json_file.write(f'\t\t"assignment_in_week": {list_of_assignment_array[i][6]},\n')
        json_file.write(f'\t\t"assignment_in_month": {list_of_assignment_array[i][7]},\n')
        
        if list_of_assignment_array[i][8] == None:
            json_file.write(f'\t\t"assignment_has_due_date": null,\n')
        
        else:
            json_file.write(f'\t\t"assignment_has_due_date": {list_of_assignment_array[i][8]},\n')
        
        json_file.write(f'\t\t"assignment_total_done": {list_of_assignment_array[i][9]},\n')

        json_file.write('\t\t"assignment_done": [')
        
        for j in range(len(list_of_assignment_array[i][10])):
            if list_of_assignment_array[i][10][j] == None:
                json_file.write('0')
            
            elif list_of_assignment_array[i][10][j] == "ü":
                json_file.write('2')
            
            elif list_of_assignment_array[i][10][j] == "NON-MUS":
                json_file.write('3')

            if j != len(list_of_assignment_array[i][10])-1:
                json_file.write(', ')

        json_file.write(']\n')
        
        if i != end_colum_range-4:
            json_file.write("\t},\n")
        
        else:
            json_file.write("\t}\n")


    json_file.write("]")
    json_file.close()


    # For Assignment Range JSON
    # Set the variables needed
    list_of_assignment_weekly_range_array = []
    for i in range(len(weekly_task_start_column_range)):
        temp_range_value = weekly_task_end_column_range[i] - weekly_task_start_column_range[i]
        list_of_assignment_weekly_range_array.append(temp_range_value)


    print(list_of_assignment_weekly_range_array)


    list_of_assignment_monthly_range_array = []
    for i in range(len(monthly_task_start_column_range)):
        temp_range_value = monthly_task_end_column_range[i] - monthly_task_start_column_range[i]
        list_of_assignment_monthly_range_array.append(temp_range_value)


    # Write to Student JSON file
    json_file = open("Excel/list_of_assignment_range.json", "w")
    json_file.write("{\n")

    json_file.write(f'\t"weekly": {list_of_assignment_weekly_range_array},\n')
    json_file.write(f'\t"monthly": {list_of_assignment_monthly_range_array}\n')

    json_file.write("}")
    json_file.close()

    


run_json_writer()

