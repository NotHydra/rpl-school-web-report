import openpyxl
from openpyxl.styles import *
from files import wb_mingguan_file, wb_bulanan_file, weekly_task_start_column_range, weekly_task_end_column_range, monthly_task_start_column_range, monthly_task_end_column_range, wb_bulanan_start_container, wb_bulanan_end_container, wb_bulanan_subtract

def write_table_names():
    wb_names = openpyxl.load_workbook("Excel/Names.xlsx")

    # Collect Data
    value_list_weekly = []
    for i in range(1, 38):
        for j in range(1,3):
            temp_value = wb_names.active.cell(row = i, column = j).value
            
            if temp_value == None:
                temp_value = " "

            value_list_weekly.append(temp_value)

    
    value_list_monthly = []
    for i in range(1, 38):
        for j in range(1,3):
            temp_value = wb_names.active.cell(row = i, column = j).value
            
            if temp_value == None:
                temp_value = " "

            value_list_monthly.append(temp_value)


    value_list_monthly.insert(2, " ")
    value_list_monthly.insert(3, " ")
            
    # Load Workbook
    wb_mingguan = []
    for i in range(len(wb_mingguan_file)):
        wb_mingguan.append(openpyxl.load_workbook(wb_mingguan_file[i]))

    wb_bulanan = []
    for i in range(len(wb_bulanan_file)):
        wb_bulanan.append(openpyxl.load_workbook(wb_bulanan_file[i]))
    

    # Write Workbook
    for wb in range(len(wb_mingguan)):
        try: 
            k = 0
            for i in range(1, 38):
                for j in range(1,3):
                    wb_mingguan[wb].active.cell(row = i, column = j, value = value_list_weekly[k])
                    k += 1

        
            wb_mingguan[wb].active.merge_cells(start_row = 1, start_column = 1, end_row = 2, end_column = 1)
            wb_mingguan[wb].active.merge_cells(start_row = 1, start_column = 2, end_row = 2, end_column = 2)

        except:
            pass

    
    for wb in range(len(wb_bulanan)):
        try: 
            k = 0
            for i in range(1, 39):
                for j in range(1,3):
                    wb_bulanan[wb].active.cell(row = i, column = j, value = value_list_monthly[k])
                    k += 1

        
            wb_bulanan[wb].active.merge_cells(start_row = 1, start_column = 1, end_row = 3, end_column = 1)
            wb_bulanan[wb].active.merge_cells(start_row = 1, start_column = 2, end_row = 3, end_column = 2)

        except:
            pass

    
    # Style
    normal = Side(border_style = "thin")
    all_border = Border(top = normal, bottom = normal, left = normal, right = normal)
    all_alignment = Alignment(horizontal ="center", vertical = "center")
    all_font = Font(size = 12)

    header_font = Font(size = 14, bold = True)

    name_alignment = Alignment(horizontal = "left", vertical = "center")

    for wb in range(len(wb_mingguan)):
        for i in range(1, 38):
            for j in range(1, 3):
                wb_mingguan[wb].active.cell(row = i, column = j).border = all_border
                wb_mingguan[wb].active.cell(row = i, column = j).alignment = all_alignment
                wb_mingguan[wb].active.cell(row = i, column = j).font = all_font


        for i in range(1, 3):
            wb_mingguan[wb].active.cell(row = 1, column = i).font = header_font  


        for i in range(3, 38):
            wb_mingguan[wb].active.cell(row = i, column = 2).alignment = name_alignment

    
    for wb in range(len(wb_bulanan)):
        for i in range(1, 39):
            for j in range(1, 3):
                wb_bulanan[wb].active.cell(row = i, column = j).border = all_border
                wb_bulanan[wb].active.cell(row = i, column = j).alignment = all_alignment
                wb_bulanan[wb].active.cell(row = i, column = j).font = all_font


        for i in range(1, 3):
            wb_bulanan[wb].active.cell(row = 1, column = i).font = header_font  


        for i in range(3, 39):
            wb_bulanan[wb].active.cell(row = i, column = 2).alignment = name_alignment


    # Save Workbook
    for i in range(len(wb_mingguan_file)):
        wb_mingguan[i].save(wb_mingguan_file[i])


    for i in range(len(wb_bulanan_file)):
        wb_bulanan[i].save(wb_bulanan_file[i])


def write_table_value():
    wb_gabunggan = openpyxl.load_workbook("Excel/List Tugas Gabunggan.xlsx")
    
    # Collect Data Value
    weekly_task_value = []
    for k in range(len(weekly_task_start_column_range)):
        value_list = []
        for j in range(weekly_task_start_column_range[k], weekly_task_end_column_range[k]):
            for i in range(2, 39):
                temp_value = wb_gabunggan.active.cell(row = i, column = j).value
                if wb_gabunggan.active.cell(row = i, column = j).value == None:
                    temp_value = ""

                value_list.append(temp_value)


        weekly_task_value.append(value_list)

    # Collect Data Value
    monthly_task_value = []
    for k in range(len(monthly_task_start_column_range)):
        value_list = []
        for j in range(monthly_task_start_column_range[k], monthly_task_end_column_range[k]):
            for i in range(1, 39):
                temp_value = wb_gabunggan.active.cell(row = i, column = j).value
                if wb_gabunggan.active.cell(row = i, column = j).value == None:
                    temp_value = ""

                value_list.append(temp_value)


        monthly_task_value.append(value_list)


    # Load Workbook
    wb_mingguan = []
    for i in range(len(wb_mingguan_file)):
        wb_mingguan.append(openpyxl.load_workbook(wb_mingguan_file[i]))


    wb_bulanan = []
    for i in range(len(wb_bulanan_file)):
        wb_bulanan.append(openpyxl.load_workbook(wb_bulanan_file[i]))


    # Write Workbook
    for wb in range(len(wb_mingguan)):
        k = 0
        for j in range(3, (weekly_task_end_column_range[wb] - weekly_task_start_column_range[wb] + 3)):
            for i in range(1, 38):
                wb_mingguan[wb].active.cell(row = i, column = j, value = weekly_task_value[wb][k])
                k += 1
    

    for wb in range(len(wb_bulanan)):
        k = 0
        for j in range(3, (monthly_task_end_column_range[wb] - monthly_task_start_column_range[wb] + 3)):
            for i in range(1, 39):
                try:
                    wb_bulanan[wb].active.cell(row = i, column = j, value = monthly_task_value[wb][k])
                
                except:
                    pass

                k += 1
    

    for wb in range(len(wb_bulanan)):
        for i in range(wb_bulanan_start_container[wb] - 1, wb_bulanan_end_container[wb] - 1): 
            if wb == 0:
                start_column = weekly_task_start_column_range[i]
                end_column = start_column
        
            elif wb == 1:
                start_column = weekly_task_start_column_range[i] - wb_bulanan_subtract[wb]
                end_column = start_column

            j = weekly_task_start_column_range[i]
            while j != weekly_task_end_column_range[i]:
                try:
                    wb_bulanan[wb].active.unmerge_cells(start_row = 1, start_column = start_column, end_row = 1, end_column = end_column)
    
                except:
                    pass
                
                end_column += 1
                j += 1


    for wb in range(len(wb_bulanan)):
        for i in range(wb_bulanan_start_container[wb] - 1, wb_bulanan_end_container[wb] - 1):
            if wb == 0:
                start_column = weekly_task_start_column_range[i]
                end_column = weekly_task_end_column_range[i] - 1
        
            elif wb == 1:
                start_column = weekly_task_start_column_range[i] - wb_bulanan_subtract[wb]
                end_column = weekly_task_end_column_range[i] - weekly_task_start_column_range[i] + 2

            try:
                wb_bulanan[wb].active.merge_cells(start_row = 1, start_column = start_column, end_row = 1, end_column = end_column)  

            except:
                pass


    # Style
    normal = Side(border_style = "thin")
    all_border = Border(top = normal, bottom = normal, left = normal, right = normal)
    all_alignment = Alignment(horizontal ="center", vertical = "center")
    all_font = Font(size = 12)

    header_font = Font(size = 14, bold = True)

    sub_header_font = Font(size = 14)

    wingdings_font = Font(name = "Wingdings", size = 12)

    light_color_fill = []
    dark_color_fill = []
    light_color = ["ffb14a", "92ff8a", "fc8dd0", "8ef5f3"]
    dark_color  = ["eb972a", "62e359", "e66cb5", "38e8e5"]
    gray_fill = PatternFill(fill_type = "solid", fgColor = "b0b0b0")

    for i in range(len(light_color)):
        light_color_fill.append(PatternFill(fill_type = "solid", fgColor = light_color[i]))
        dark_color_fill.append(PatternFill(fill_type = "solid", fgColor = dark_color[i]))
    

    for wb in range(len(wb_mingguan)):
        for i in range(1, 38):
            for j in range(3, (weekly_task_end_column_range[wb] - weekly_task_start_column_range[wb] + 3)):
                wb_mingguan[wb].active.cell(row = i, column = j).border = all_border
                wb_mingguan[wb].active.cell(row = i, column = j).alignment = all_alignment
                wb_mingguan[wb].active.cell(row = i, column = j).font = all_font


        for i in range(3, (weekly_task_end_column_range[wb] - weekly_task_start_column_range[wb] + 3)):
            wb_mingguan[wb].active.cell(row = 1, column = i).font = header_font


        for i in range(3, (weekly_task_end_column_range[wb] - weekly_task_start_column_range[wb] + 3)):
            wb_mingguan[wb].active.cell(row = 2, column = i).font = sub_header_font


        for i in range(3, 38):
            for j in range(3, (weekly_task_end_column_range[wb] - weekly_task_start_column_range[wb] + 3)):
                if wb_mingguan[wb].active.cell(row = i, column = j).value == "ü":
                    wb_mingguan[wb].active.cell(row = i, column = j).font = wingdings_font
                
                if j % 2 == 0:
                    wb_mingguan[wb].active.cell(row = i, column = j).fill = gray_fill

    
    for wb in range(len(wb_bulanan)):
        for i in range(1, 39):
            for j in range(3, (monthly_task_end_column_range[wb] - monthly_task_start_column_range[wb] + 3)):
                wb_bulanan[wb].active.cell(row = i, column = j).border = all_border
                wb_bulanan[wb].active.cell(row = i, column = j).alignment = all_alignment
                wb_bulanan[wb].active.cell(row = i, column = j).font = all_font


        for i in range(3, (monthly_task_end_column_range[wb] - monthly_task_start_column_range[wb] + 3)):
            wb_bulanan[wb].active.cell(row = 1, column = i).font = header_font


        for i in range(3, (monthly_task_end_column_range[wb] - monthly_task_start_column_range[wb] + 3)):
            wb_bulanan[wb].active.cell(row = 2, column = i).font = sub_header_font


        for i in range(3, 39):
            for j in range(3, (monthly_task_end_column_range[wb] - monthly_task_start_column_range[wb] + 3)):
                if wb_bulanan[wb].active.cell(row = i, column = j).value == "ü":
                    wb_bulanan[wb].active.cell(row = i, column = j).font = wingdings_font
        

        l = 0
        for k in range(wb_bulanan_start_container[wb] - 1, wb_bulanan_end_container[wb] - 1):
            if wb == 0:
                start_column = weekly_task_start_column_range[k]
                end_column = weekly_task_end_column_range[k]

            elif wb == 1:
                start_column = weekly_task_start_column_range[k] - wb_bulanan_subtract[wb]
                end_column = weekly_task_end_column_range[k] - weekly_task_start_column_range[k] + 3

            for i in range(1, 39):
                for j in range(start_column, end_column):
                    wb_bulanan[wb].active.cell(row = i, column = j).fill = light_color_fill[l]  
            l += 1

            if l >= 4:
                l = 0 
        

        l = 0
        for k in range(wb_bulanan_start_container[wb] - 1, wb_bulanan_end_container[wb] - 1):
            if wb == 0:
                start_column = weekly_task_start_column_range[k]
                end_column = weekly_task_end_column_range[k]
            
            elif wb == 1:
                start_column = weekly_task_start_column_range[k] - wb_bulanan_subtract[wb]
                end_column = weekly_task_end_column_range[k] - weekly_task_start_column_range[k] + 3

            for i in range(4, 39):
                for j in range(start_column, end_column):
                    if j % 2 == 0:
                        wb_bulanan[wb].active.cell(row = i, column = j).fill = dark_color_fill[l]
            l += 1

            if l >= 4 :
                l = 0


    # Save Workbook
    for i in range(len(wb_mingguan_file)):
        wb_mingguan[i].save(wb_mingguan_file[i])
    
    for i in range(len(wb_bulanan_file)):
        wb_bulanan[i].save(wb_bulanan_file[i])


def check_validity():
    # Load Workbook
    wb_gabunggan = openpyxl.load_workbook("Excel/List Tugas Gabunggan.xlsx")

    wb_bulanan = []
    for i in range(len(wb_bulanan_file)):
        wb_bulanan.append(openpyxl.load_workbook(wb_bulanan_file[i]))


    wb_mingguan = []
    for i in range(len(wb_mingguan_file)):
        wb_mingguan.append(openpyxl.load_workbook(wb_mingguan_file[i]))


    # Task Value
    combined_task_value = []
    for k in range(len(weekly_task_start_column_range)):
        for i in range(4, 39):
            for j in range(weekly_task_start_column_range[k], weekly_task_end_column_range[k]):
                temp_value = wb_gabunggan.active.cell(row = i, column = j).value

                if temp_value == None:
                    temp_value = ""
                
                combined_task_value.append(temp_value)


    combined_monthly_task_value = []
    for wb in range(len(wb_bulanan)):
        temp_value_list = []
        for i in range(4, 39):
            for j in range(monthly_task_start_column_range[wb], monthly_task_end_column_range[wb]):
                temp_value = wb_gabunggan.active.cell(row = i, column = j).value

                if temp_value == None:
                    temp_value = ""

                temp_value_list.append(temp_value)

        combined_monthly_task_value.append(temp_value_list)


    combined_weekly_task_value = []
    for wb in range(len(wb_mingguan)):
        temp_value_list = []
        for i in range(4, 39):
            for j in range(weekly_task_start_column_range[wb], weekly_task_end_column_range[wb]):
                temp_value = wb_gabunggan.active.cell(row = i, column = j).value

                if temp_value == None:
                    temp_value = ""

                temp_value_list.append(temp_value)

        combined_weekly_task_value.append(temp_value_list)
    

    monthly_task_value = []
    for wb in range(len(wb_bulanan)):
        temp_value_list = []
        for i in range(4, 39):
            for j in range(3, monthly_task_end_column_range[wb] - monthly_task_start_column_range[wb] + 3):
                temp_value = wb_bulanan[wb].active.cell(row = i, column = j).value

                if temp_value == None:
                    temp_value = ""
                
                temp_value_list.append(temp_value)
        

        monthly_task_value.append(temp_value_list)


    weekly_task_value = []
    for wb in range(len(wb_mingguan)):
        temp_value_list = []
        for i in range(3, 38):
            for j in range(3, weekly_task_end_column_range[wb] - weekly_task_start_column_range[wb] + 3):
                temp_value = wb_mingguan[wb].active.cell(row = i, column = j).value

                if temp_value == None:
                    temp_value = ""

                temp_value_list.append(temp_value)


        weekly_task_value.append(temp_value_list)


    # Print Max Value
    # Print Combined Max Value
    print()
    print("Checking Validity")
    print(f"Combined Max Value : {len(combined_task_value)},", end = " ")

    total_value = 0
    for i in range(len(combined_monthly_task_value)):
        total_value += len(combined_monthly_task_value[i])


    print(f"{total_value},", end = " ")

    total_value = 0
    for i in range(len(combined_weekly_task_value)):
        total_value += len(combined_weekly_task_value[i])


    print(f"{total_value}")
    print()

    # Print Month Max Value
    print(f"C.Month Max Value :", end = " ")
    for i in range(len(combined_monthly_task_value)):
        if i != len(combined_monthly_task_value) - 1:
            print(f"{len(combined_monthly_task_value[i])},", end = " ")

        elif i == len(combined_monthly_task_value) - 1:
            print(f"{len(combined_monthly_task_value[i])}")
    
    
    print(f"Month Max Value   :", end = " ")
    for i in range(len(monthly_task_value)):
        if i != len(monthly_task_value) - 1:
            print(f"{len(monthly_task_value[i])},", end = " ")

        elif i == len(monthly_task_value) - 1:
            print(f"{len(monthly_task_value[i])}")
    

    # Print Weekly Max Value
    print()
    print(f"C.Weekly Max Value :", end = " ")
    for i in range(len(combined_weekly_task_value)):
        if i != len(combined_weekly_task_value) - 1:
            print(f"{len(combined_weekly_task_value[i])},", end = " ")

        elif i == len(combined_weekly_task_value) - 1:
            print(f"{len(combined_weekly_task_value[i])}")


    print(f"Weekly Max Value   :", end = " ")
    for i in range(len(weekly_task_value)):
        if i != len(weekly_task_value) - 1:
            print(f"{len(weekly_task_value[i])},", end = " ")

        elif i == len(weekly_task_value) - 1:
            print(f"{len(weekly_task_value[i])}")
    

    # Print Current Value
    # Combined Current Value
    print()
    print()
    print()
    combined_task_current_value = 0
    for i in range(len(combined_task_value)):
        if combined_task_value[i] != "":
            combined_task_current_value += 1


    print(f"Combined Current Value : {combined_task_current_value},", end = " ")

    combined_monthly_task_current_value = 0
    for wb in range(len(wb_bulanan)):
        for i in range(len(combined_monthly_task_value[wb])):
            if combined_monthly_task_value[wb][i] != "":
                combined_monthly_task_current_value += 1


    print(f"{combined_monthly_task_current_value},", end = " ")

    combined_weekly_task_current_value = 0
    for wb in range(len(wb_mingguan)):
        for i in range(len(combined_weekly_task_value[wb])):
            if combined_weekly_task_value[wb][i] != "":
                combined_weekly_task_current_value += 1


    print(f"{combined_weekly_task_current_value}")
    print()

    # Monthly Current Value
    print(f"C.Month Current Value :", end = " ")
    for wb in range(len(wb_bulanan)):
        combined_monthly_task_current_value = 0
        for i in range(len(combined_monthly_task_value[wb])):
            if combined_monthly_task_value[wb][i] != "":
                combined_monthly_task_current_value += 1

        if wb != len(wb_bulanan) - 1:
            print(f"{combined_monthly_task_current_value},", end = " ")
        
        elif wb == len(wb_bulanan) - 1:
            print(f"{combined_monthly_task_current_value}")
        
    
    print(f"Month Current Value   :", end = " ")
    for wb in range(len(wb_bulanan)):
        monthly_task_current_value = 0
        for i in range(len(monthly_task_value[wb])):
            if monthly_task_value[wb][i] != "":
                monthly_task_current_value += 1


        if wb != len(wb_bulanan) - 1:
            print(f"{monthly_task_current_value},", end = " ")
        
        elif wb == len(wb_bulanan) - 1:
            print(f"{monthly_task_current_value}")
        
    
    print()
    
    # Weekly Current Value
    print(f"C.Weekly Current Value :", end = " ")
    for wb in range(len(wb_mingguan)):
        combined_weekly_task_current_value = 0
        for i in range(len(combined_weekly_task_value[wb])):
            if combined_weekly_task_value[wb][i] != "":
                combined_weekly_task_current_value += 1


        if wb != len(wb_mingguan) - 1:
            print(f"{combined_weekly_task_current_value},", end = " ")

        elif wb == len(wb_mingguan) - 1:
            print(f"{combined_weekly_task_current_value}")


    print(f"Weekly Current Value   :", end = " ")
    for wb in range(len(wb_mingguan)):
        weekly_task_current_value = 0
        for i in range(len(weekly_task_value[wb])):
            if weekly_task_value[wb][i] != "":
                weekly_task_current_value += 1

        
        if wb != len(wb_mingguan) - 1:
            print(f"{weekly_task_current_value},", end = " ")

        elif wb == len(wb_mingguan) - 1:
            print(f"{weekly_task_current_value}")


    print()
        

write_table_names()
write_table_value()
check_validity()