import openpyxl
from openpyxl.styles import *

def write_table_names():
    wb_names = openpyxl.load_workbook("Excel/Names.xlsx")

    # Collect Data
    value_list = []
    for i in range(1, 38):
        for j in range(1,3):
            temp_value = wb_names.active.cell(row = i, column = j).value
            
            if temp_value == None:
                temp_value = " "

            value_list.append(temp_value)

    # Load Workbook + Update
    wb_mingguan = []
    wb_mingguan.append(openpyxl.load_workbook("Excel/Mingguan/List Tugas Minggu Ke-1 (26 Juli 2021 -  1 Agustus 2021).xlsx"))
    wb_mingguan.append(openpyxl.load_workbook("Excel/Mingguan/List Tugas Minggu Ke-2 (2 Agustus 2021 - 8 Agustus 2021).xlsx"))
    wb_mingguan.append(openpyxl.load_workbook("Excel/Mingguan/List Tugas Minggu Ke-3 (9 Agustus 2021 - 15 Agustus 2021).xlsx"))
    wb_mingguan.append(openpyxl.load_workbook("Excel/Mingguan/List Tugas Minggu Ke-4 (16 Agustus 2021 - 22 Agustus 2021).xlsx"))
    wb_mingguan.append(openpyxl.load_workbook("Excel/Mingguan/List Tugas Minggu Ke-5 (23 Agustus 2021 - 29 Agustus 2021).xlsx"))
    
    # Write Workbook
    for wb in range(len(wb_mingguan)):
        try: 
            k = 0
            for i in range(1, 38):
                for j in range(1,3):
                    wb_mingguan[wb].active.cell(row = i, column = j, value = value_list[k])
                    k += 1

        
            wb_mingguan[wb].active.merge_cells(start_row = 1, start_column = 1, end_row = 2, end_column = 1)
            wb_mingguan[wb].active.merge_cells(start_row = 1, start_column = 2, end_row = 2, end_column = 2)

        except:
            pass

    
    # Style
    normal = Side(border_style = "thin")
    all_border = Border(top = normal, bottom = normal, left = normal, right = normal)
    all_alignment = Alignment(horizontal ="center", vertical = "center")
    all_font = Font(size = 12)

    header_font = Font(size = 14, bold = True)

    name_alignment = Alignment(horizontal = "left", vertical = "center")

    for wb in range(len(wb_mingguan)):
        for i in range(1, 38):
            for j in range(1, 3):
                wb_mingguan[wb].active.cell(row = i, column = j).border = all_border
                wb_mingguan[wb].active.cell(row = i, column = j).alignment = all_alignment
                wb_mingguan[wb].active.cell(row = i, column = j).font = all_font


        for i in range(1, 3):
            wb_mingguan[wb].active.cell(row = 1, column = i).font = header_font  


        for i in range(3, 38):
            wb_mingguan[wb].active.cell(row = i, column = 2).alignment = name_alignment


    # Save Workbook + Update
    wb_mingguan[0].save("Excel/Mingguan/List Tugas Minggu Ke-1 (26 Juli 2021 -  1 Agustus 2021).xlsx")
    wb_mingguan[1].save("Excel/Mingguan/List Tugas Minggu Ke-2 (2 Agustus 2021 - 8 Agustus 2021).xlsx")
    wb_mingguan[2].save("Excel/Mingguan/List Tugas Minggu Ke-3 (9 Agustus 2021 - 15 Agustus 2021).xlsx")
    wb_mingguan[3].save("Excel/Mingguan/List Tugas Minggu Ke-4 (16 Agustus 2021 - 22 Agustus 2021).xlsx")
    wb_mingguan[4].save("Excel/Mingguan/List Tugas Minggu Ke-5 (23 Agustus 2021 - 29 Agustus 2021).xlsx")


def write_table_value():
    wb_gabunggan = openpyxl.load_workbook("Excel/List Tugas Gabunggan.xlsx")
    
    # Collect Data Value + Update
    weekly_task_start_column_range = [3, 8, 16, 23, 26]
    weekly_task_end_column_range   = [8, 16, 23, 26, 27]
    weekly_task_value = []
    for k in range(len(weekly_task_start_column_range)):
        value_list = []
        for j in range(weekly_task_start_column_range[k], weekly_task_end_column_range[k]):
            for i in range(2, 39):
                temp_value = wb_gabunggan.active.cell(row = i, column = j).value
                if wb_gabunggan.active.cell(row = i, column = j).value == None:
                    temp_value = ""

                value_list.append(temp_value)


        weekly_task_value.append(value_list)


    """
    # Print Out Data Value Per Week
    for i in range(len(weekly_task_value)):
        print(weekly_task_value[i])
        print(len(weekly_task_value[i]))
        print()
    """


    # Load Workbook  + Update
    wb_mingguan = []
    wb_mingguan.append(openpyxl.load_workbook("Excel/Mingguan/List Tugas Minggu Ke-1 (26 Juli 2021 -  1 Agustus 2021).xlsx"))
    wb_mingguan.append(openpyxl.load_workbook("Excel/Mingguan/List Tugas Minggu Ke-2 (2 Agustus 2021 - 8 Agustus 2021).xlsx"))
    wb_mingguan.append(openpyxl.load_workbook("Excel/Mingguan/List Tugas Minggu Ke-3 (9 Agustus 2021 - 15 Agustus 2021).xlsx"))
    wb_mingguan.append(openpyxl.load_workbook("Excel/Mingguan/List Tugas Minggu Ke-4 (16 Agustus 2021 - 22 Agustus 2021).xlsx"))
    wb_mingguan.append(openpyxl.load_workbook("Excel/Mingguan/List Tugas Minggu Ke-5 (23 Agustus 2021 - 29 Agustus 2021).xlsx"))

    # Write Workbook
    for wb in range(len(wb_mingguan)):
        k = 0
        for j in range(3, (weekly_task_end_column_range[wb] - weekly_task_start_column_range[wb] + 3)):
            for i in range(1, 38):
                wb_mingguan[wb].active.cell(row = i, column = j, value = weekly_task_value[wb][k])
                k += 1


    # Style
    normal = Side(border_style = "thin")
    all_border = Border(top = normal, bottom = normal, left = normal, right = normal)
    all_alignment = Alignment(horizontal ="center", vertical = "center")
    all_font = Font(size = 12)

    header_font = Font(size = 14, bold = True)

    sub_header_font = Font(size = 14)

    wingdings_font = Font(name = "Wingdings", size = 12)

    gray_fill = PatternFill(fill_type = "solid", fgColor = "b0b0b0")

    for wb in range(len(wb_mingguan)):
        for i in range(1, 38):
            for j in range(3, (weekly_task_end_column_range[wb] - weekly_task_start_column_range[wb] + 3)):
                wb_mingguan[wb].active.cell(row = i, column = j).border = all_border
                wb_mingguan[wb].active.cell(row = i, column = j).alignment = all_alignment
                wb_mingguan[wb].active.cell(row = i, column = j).font = all_font


        for i in range(3, (weekly_task_end_column_range[wb] - weekly_task_start_column_range[wb] + 3)):
            wb_mingguan[wb].active.cell(row = 1, column = i).font = header_font


        for i in range(3, (weekly_task_end_column_range[wb] - weekly_task_start_column_range[wb] + 3)):
            wb_mingguan[wb].active.cell(row = 2, column = i).font = sub_header_font


        for i in range(3, 38):
            for j in range(3, (weekly_task_end_column_range[wb] - weekly_task_start_column_range[wb] + 3)):
                if wb_mingguan[wb].active.cell(row = i, column = j).value == "ü":
                    wb_mingguan[wb].active.cell(row = i, column = j).font = wingdings_font
                
                if j % 2 == 0:
                    wb_mingguan[wb].active.cell(row = i, column = j).fill = gray_fill


    # Save Workbook + Update
    wb_mingguan[0].save("Excel/Mingguan/List Tugas Minggu Ke-1 (26 Juli 2021 -  1 Agustus 2021).xlsx")
    wb_mingguan[1].save("Excel/Mingguan/List Tugas Minggu Ke-2 (2 Agustus 2021 - 8 Agustus 2021).xlsx")
    wb_mingguan[2].save("Excel/Mingguan/List Tugas Minggu Ke-3 (9 Agustus 2021 - 15 Agustus 2021).xlsx")
    wb_mingguan[3].save("Excel/Mingguan/List Tugas Minggu Ke-4 (16 Agustus 2021 - 22 Agustus 2021).xlsx")
    wb_mingguan[4].save("Excel/Mingguan/List Tugas Minggu Ke-5 (23 Agustus 2021 - 29 Agustus 2021).xlsx")

write_table_names()
write_table_value()